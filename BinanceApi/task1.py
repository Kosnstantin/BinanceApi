import sqlite3
from binance import Client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import subprocess


# symbol = input("Input your symbol (e.g., BTCUSDT, ETHUSDT, etc): ")
# interval = input("Input your interval (e.g., 1d, 4h, 1h): ")
apikey = ''
secret = ''


client = Client(apikey, secret)
# historical = client.get_historical_klines(symbol, interval, '1 June 2023')
historical = client.get_historical_klines('ETHUSDT', Client.KLINE_INTERVAL_1HOUR, '1 June 2023')

data = {
    'Open time': [],
    'Open price': [],
    'The highest price': [],
    'The lowest price': [],
    'Close price': [],
    'Volume of trades': [],
    'Close time': [],
}
connect = sqlite3.connect('klines.db')
cursor = connect.cursor()
query = """CREATE TABLE IF NOT EXISTS klines(Open_time REAL, Open_price REAL, The_highest_price REAL, The_lowest_price REAL, Close_price REAL, Volume_of_trades REAL, Close_time REAL)"""

cursor.execute(query)
query1 = """INSERT INTO klines (Open_time, Open_price, The_highest_price, The_lowest_price, Close_price, Volume_of_trades, Close_time) VALUES(?, ?, ?, ?, ?, ?, ?)"""

for klines in historical:
    time_open = klines[0]
    price_open = klines[1]
    top_price = klines[2]
    low_price = klines[3]
    price_close = klines[4]
    trading_volume = klines[5]
    time_close = klines[6]

    data['Open time'].append(time_open)
    data['Open price'].append(price_open)
    data['The highest price'].append(top_price)
    data['The lowest price'].append(low_price)
    data['Close price'].append(price_close)
    data['Volume of trades'].append(trading_volume)
    data['Close time'].append(time_close)

    cursor.execute(query1, (
        time_open, price_open, top_price, low_price, price_close, trading_volume, time_close
    ))  # Insert db

# DataFrame is a data structure (table with indexes and columns)
df = pd.DataFrame(data)

filename = 'klines.csv'


wb = Workbook()  # Excel book
ws = wb.active  # page for data

for col_num, col_name in enumerate(df.columns, start=1):  # Column enumeration

    ws.cell(row=1, column=col_num, value=col_name)  # Columns name

    ws.cell(row=1, column=col_num).font = Font(bold=True)

    ws.cell(row=1, column=col_num).alignment = Alignment(
        horizontal='center')  # Placement of text

    for row_num, value in enumerate(df[col_name], start=2):

        cell = ws.cell(row=row_num, column=col_num,
                       value=value)  # Insert text in cells
        cell.alignment = Alignment(horizontal='center')  # Placement of text

    column_letter = chr(64 + col_num)  # Get letter of columns (A-Z)
    # Info about the size and properties
    column_dimensions = ws.column_dimensions[column_letter]
    column_dimensions.width = 20


wb.save(filename)


connect.commit()  # Make all changes
connect.close()

print(f"Data saved to a file {filename} and a database")
